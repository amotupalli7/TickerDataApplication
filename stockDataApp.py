# This script contains the code of a automatic stock data tracker which will be able to run in an executable

# Packages needed: requests, openpxyl, datetime, math
#############################################################################################################################
from re import I
from openpyxl import load_workbook
import requests, openpyxl, datetime, math, time
from datetime import date, datetime, timedelta
from openpyxl import workbook
from tkinter import *
from tkinter.filedialog import askopenfilename 
#############################################################################################################################


#############################################################################################################################
# Functions to format data correctly for api calls and inputs
#############################################################################################################################

def dateToMili(dtObj): return int(datetime.timestamp(dtObj)*1000) # (year,month,day): miliseconds since epoch
def strToDate(str): return datetime.strptime(str,'%m/%d/%Y') # (MM/DD/YYYY): datetimeObject
def strToMili(str): return dateToMili(strToDate(str)) # (MM/DD/YYYY): miliseconds sinch epoch

#############################################################################################################################
# Makes a request to TD Ameritrade's API and returns price history based on the arguments passed
# Refer to TD Ameritrade market data API for proper arguments
#  Ticker	Date    Gap % Before Open	Open 	High    Low    Close	Volume(M)	Float(M)	Market Cap(M)   Float Rotation	Max % Gain	Day % Change    Day 2 Open  Day2 High   Day2 Low    Day2 Close  
#############################################################################################################################

def makeRequest(**kwargs):
    url = "https://api.tdameritrade.com/v1/marketdata/{}/pricehistory".format(kwargs.get('symbol'))
    apiKey = "ABCHASFPBYFGHJM1OG6GSH3GU23UPT8C" # Varies by User, before publically publishing this program, add a feature for this to given by user
    parameters = {}
    parameters.update({'apikey': apiKey})

    for arg in kwargs:  
        p = {arg: kwargs.get(arg)}
        parameters.update(p)    
    
    return requests.get(url, params=parameters).json()

#############################################################################################################################
# Gets and formats ticker data using makeRequest function
#############################################################################################################################

def getTickerData(ticker, sDate, eDate): # (Ticker, datetime obj, datetime obj)

    startDate = dateToMili(sDate)
    endDate = dateToMili(eDate) 
    data = makeRequest(symbol= ticker, periodType = 'year', frequency= 1, frequencyType = 'daily', startDate= startDate, endDate = endDate)
    time.sleep(0.5)
    data2 = data["candles"]
    for day in data2:
        day["volume"] = int(day["volume"] / 1000000)
        day["datetime"] = date.fromtimestamp(int(day["datetime"]/1000))
        dtString = day["datetime"].strftime("%m/%d/%Y")
        day["datetime"] = dtString
    
    return data2

#############################################################################################################################
# Function that gets the gapping percentage from the date given compared to the previous day's close
#############################################################################################################################

def getGapPercentage(ticker,dt): # (String: 'AAPL', datetime)
    
    lastDate = dt - timedelta(days=5)

    if(dt > datetime.today()):
        raise Exception("Date provided can't be in the future")
    # if(dt.isoweekday()==1): # if the day is a monday
    #     lastDate = dt- timedelta(days=3)
    # else: lastDate = dt - timedelta(days=1)

    data = getTickerData(ticker,lastDate,dt)
    if(len(data)==0): # The request is bad if its still one after giving a date
        print(f"Bad Request for {ticker} Gap Percentage.")
        return None

    # while(len(data)!=2):
    #     lastDate = lastDate - timedelta(days=1)
    #     data = getTickerData(ticker,lastDate,dt)

    
    cPrice = data[len(data)-2]["close"]
    oPrice = data[len(data)-1]["open"]
    return math.ceil(((oPrice-cPrice)/cPrice) *100)  



#############################################################################################################################
# Function that provides rawData to fill into each row
#############################################################################################################################
def fillRow(ticker, date, numDays): # Takes in ticker, dateObj, and number of days you want data for | returns a list of data containing all data 
    data = []
    def helper(ticker,eDate,numDays):
        # check if numDays is going to cause an error
        if(eDate  > datetime.today()):
            print(f"Number of days is too high for ticker {ticker} on {date}")
            return []
        # Otherwise makes call to request function and formats data, return has to include d
        data = getTickerData(ticker,date,eDate)
        if(len(data)==0):
            print(f"Bad Request for ticker {ticker}")
            return data
        while len(data) < numDays: # This means the end date is not a market day
            # If the end date is a friday then recursively call the method after adding 3 days else just add one day and recursively call
            return helper(ticker,eDate + timedelta(days=3),numDays) if (endDate.isoweekday()==5) else helper(ticker,eDate + timedelta(days=1),numDays)
        return data

    endDate = date + timedelta(days=numDays-1) 
    # check if numDays is going to cause an error
    if(endDate  > datetime.today()):
        print(f"Number of days is too high for ticker {ticker} on {date}")
    else:
        data = helper(ticker,endDate,numDays)
    output = [ticker,date,getGapPercentage(ticker,date)]
    for item in data:
        output.extend([item['open'],item['high'],item['low'],item['close'],item['volume']])
    return output
    

 
#############################################################################################################################
# Function that fills empty cells in an excel file given the ticker and date
#############################################################################################################################
def fillRawData(inputFile, page): # Takes in an excel file and a sheet in that file | no return just fills in the cells in that excel file

    wb = openpyxl.load_workbook(inputFile)

    ws = wb[page]
    headers = list(ws['1'])
    numDays = int((len(headers) - 3)/5) # The header will always be ticker, date, gap Percentage and then for each day they want data, there will be 5 elements ( Open, High, Low, Close, Volume )
    print("Data filled for: ")
    for row in ws.iter_rows(min_row=2): # Starting at the 2nd row because the first is the headers
        
        if row[len(headers)-1].value == None and row[0].value !=None: # Only fill data for rows that have an empty last cell
            ticker = row[0].value
            date = row[1].value # Have to check what type of data this is
            data = fillRow(ticker,date,numDays)

            for index in range(0,len(data)):
                row[index].value = data[index]
            print(ticker)

    wb.save(filename=inputFile)

############################################################################################################################################################
# Gets minute data for the specified date for specified ticker
############################################################################################################################################################

def getMinuteData(ticker,date):
    symbol = ticker
    sDate = dateToMili(date)
    periodType = 'day'
    frequencyType = 'minute'
    eDate = dateToMili(date + timedelta(days=1))
    needExtendedHoursData = 'true'
    data = makeRequest(symbol=symbol,periodType=periodType,frequencyType=frequencyType,startDate=sDate,endDate=eDate, needExtendedHoursData=needExtendedHoursData)
    data = data["candles"]
    for day in data:
        day["datetime"] = datetime.fromtimestamp(int(day["datetime"]/1000))
        dtString = day["datetime"].strftime("%H:%M:%S")
        day["datetime"] = dtString
    
    return data

############################################################################################################################################################
# Gets regular market time data
############################################################################################################################################################
def getMarketHourData(ticker,date):
    data = getMinuteData(ticker,date)
    marketTimeData = []
    bool = False
    for day in data:
        if(day['datetime'] == '09:30:00'): bool= True
        if(day['datetime'] == '16:01:00'): bool = False
        if(bool): marketTimeData.append(day)

    return marketTimeData


        

############################################################################################################################################################
# Gets premarket minute data
############################################################################################################################################################

def getPremarketData(ticker,date):
    minuteData = getMinuteData(ticker,date)
    premarketData = []
    for day in minuteData:
        if(day['datetime'] == '09:30:00'): return premarketData
        premarketData.append(day)
    
    return(premarketData)

############################################################################################################################################################
# Gets premarket high information 
############################################################################################################################################################

def getPremarketHighInfo(ticker,date):
    data = getPremarketData(ticker,date)
    premarketHigh = 0
    premarketHighTime = ""
    for minute in data:
        if minute['high'] >= premarketHigh: 
            premarketHigh = minute['high']
            premarketHighTime = minute['datetime']
    
    return (premarketHigh,premarketHighTime)

############################################################################################################################################################
# Gets HOD and LOD time information
############################################################################################################################################################

def getRegularHODandLODTime(ticker,date):
    data = getMarketHourData(ticker,date)
    HOD = 0
    LOD = 1000000000000000
    HighTime = ""
    LowTime = ""
    for minute in data:
        if minute['high'] >= HOD:
            HOD = minute['high']
            HighTime = minute['datetime']
        if minute['low'] <= LOD:
            LOD = minute['low']
            LowTime = minute['datetime']
    return(HighTime,LowTime)

#############################################################################################################################
# Updated fill raw data with premarket high, each day HOD and LOD time just day 1 and two
#############################################################################################################################
def updatedFillRawData(inputFile, page): # assumes the headers are ticker, date, gap % , premarket high, premarket high time, HOD time, LOD time (open, high, low, close, volume)


    wb = openpyxl.load_workbook(inputFile)

    ws = wb[page]
    headers = list(ws['1'])
    numDays = int((len(headers) - 7)/5) # The header will always be ticker, date, gap Percentage and then for each day they want data, there will be 5 elements ( Open, High, Low, Close, Volume )
    print("Data filled for: ")
    for row in ws.iter_rows(min_row=2): # Starting at the 2nd row because the first is the headers
        
        if row[len(headers)-1].value == None and row[0].value !=None: # Only fill data for rows that have an empty last cell
            ticker = row[0].value
            date = row[1].value # Have to check what type of data this is
            data = fillRow(ticker,date,numDays)


            premarketHighInfo = (getPremarketHighInfo(ticker,date)) # (1.43, '09:18:00')
            timeInfo = (getRegularHODandLODTime(ticker,date)) #('09:30:00','10:00:00')
            
            data[3:3] = premarketHighInfo
            data[5:5] = timeInfo
            for index in range(0,len(data)):
                row[index].value = data[index]
            print(ticker)

    wb.save(filename=inputFile)

    

        


#############################################################################################################################
# Main method
#############################################################################################################################

if __name__ =='__main__':

    

    

    #print(strToMili("06/26/2022"))
    #print(getTickerData("AAPL",strToDate('06/22/2022'),strToDate('06/22/2022')))
    #print(getGapPercentage('BHAT','06/24/2022'))
    #print(getGapPercentage('AAPL','06/29/2022'))
    sheet = input("Enter excel file path which stock data would like to be updated ")
    #sheet = 'DailyGapperRawData.xlsx'     
    #fillRawData(sheet)

    page = input("Enter sheet in excel file: ")

    fillRawData(sheet,page)
    #updatedFillRawData(sheet,page)

    input("Press ENTER to close script. ")

    #print(fillRow("HYRE", strToDate("09/26/2022"),2))

    
    


    